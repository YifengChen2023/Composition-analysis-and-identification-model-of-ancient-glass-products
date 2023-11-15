import copy
import numpy as np
import math
import csv
from random import random
import openpyxl
import matplotlib.pyplot as plt

def loadDataset(filename): # 读入数据
    with open(filename, 'r') as f:
        lines = csv.reader(f)
        data_set = list(lines)

    # 整理数据
    feature = data_set[0]
    del (data_set[0])

    for i in range(len(data_set)):
        for j in range(len(data_set[0])):
            data_set[i][j] = float(data_set[i][j])

    return data_set, feature

def get_dist(p1, p2): # 计算点间的欧氏距离
    dist_ = 0
    for i in range(len(p1)):
        dist_ += (p1[i] - p2[i]) ** 2
    return math.sqrt(dist_)

def nearest(cur_point, cluster_centers): # 找到距离当前节点最近的聚类中心点
    min_dist = np.inf
    m = np.shape(cluster_centers)[0]  # 当前已经初始化的聚类中心的个数
    for i in range(m):
        d = get_dist(cur_point, cluster_centers[i])  # 计算cur_point与每个聚类中心之间的距离
        if min_dist > d:  # 选择最短距离
            min_dist = d
    return min_dist

def get_centroids(data, k): # K-means++算法，产生初始的聚类中心
    data_set = copy.deepcopy(data)
    m = len(data_set)  # 行数 即数据点的个数
    n = len(data_set[0])  #列数 即特征的个数

    cluster_centers = []  # 初始化聚类中心
    for i in range(k):
        cluster_centers.append([])

    index = np.random.randint(0, m)  # 1、随机选择一个样本点为第一个聚类中心
    cluster_centers[0] = copy.deepcopy(data_set[index])

    d = [0.0 for _ in range(m)]  # 2、初始化一个距离的序列
    for i in range(1, k): # 外层是更新cluster_centers的值
        sum_all = 0
        for j in range(m):
            d[j] = nearest(data_set[j], cluster_centers[0:i])  # 3、对每一个样本找到最近的聚类中心点 并返回最近距离值
            sum_all += d[j]  # 4、将所有的最短距离相加
        sum_all *= random()  # 5、取得sum_all之间的随机值
        for j, di in enumerate(d):  # 6、获得距离最远的样本点作为聚类中心点
            sum_all -= di
            if sum_all > 0:
                continue
            cluster_centers[i] = copy.deepcopy(data_set[j])
            break
    return cluster_centers

def generate_clusters(k, data, cluster_centers, n): # k-means算法 更新迭代聚类中心
    flag = 1
    while flag:
        flag = 0
        cluster = []
        for i in range(k):
            cluster.append([])
        for i in range(len(data)):
            d = float('inf')
            for j in range(k):
                d_j = get_dist(data[i], cluster_centers[j])
                if d > d_j:
                    d = d_j
                    center_num = j
            cluster[center_num].append(i)
        for i in range(k):
            for j in range(len(cluster[i])):
                if j == 0:
                    new_center = copy.deepcopy(data[cluster[i][0]])
                    continue
                for q in range(n):
                    new_center[q] += data[cluster[i][j]][q]
            for p in range(len(new_center)):
                new_center[p] /= len(cluster[i])
            if new_center != cluster_centers[i]:
                cluster_centers[i] = new_center
                flag = 1

    return cluster_centers, cluster

def get_sse(k, cluster_centers, clusters, data): # 计算聚类划分评价值和方差SSE
    sum_see = 0
    for i in range(k):
        for x in clusters[i]:
            sum_see += get_dist(data[x], cluster_centers[i]) ** 2
    return sum_see

def get_Silhouette_Coefficient(k, cluster, data): # 计算聚类划分评价值轮廓系数
    if k == 1:
        return 0
    else:
        sum_s_co = np.zeros(len(data))
        for i in range(len(data)):
            a = 0

            for cluster_instance in cluster:
                if i in cluster_instance:
                    cluster_instance_i = cluster_instance
                    break
            for j in cluster_instance_i:
                 a += get_dist(data[i], data[j])
            a /= len(cluster_instance_i)
            b_min = float('inf')
            for cluster_instance in cluster:
                b = 0
                if cluster_instance != cluster_instance_i:
                    for j in cluster_instance:
                        b += get_dist(data[i], data[j])
                    b /= len(cluster_instance)

                    if b < b_min:
                        b_min = b
            sum_s_co[i] = (b_min - a) / max(b_min, a)
        return sum_s_co.mean()

def get_similarity(base_cluster, cur_cluster): # 计算模型稳定行和敏感性指标 划分相似度
    mean_similarity = 0
    for c in cur_cluster:
        jaccard = 0
        for b in base_cluster:
            c = set(c)
            b = set(b)
            tmp = len(c&b)/len(c|b)
            if tmp > jaccard:
                jaccard = tmp
        mean_similarity += jaccard
    return mean_similarity/4

if __name__ == '__main__':

    # 仅以计算相似度为例展示函数的调用关系 其他main函数不再赘述

    data_set, feature_list = loadDataset('k-means-PbBa.csv')
    wb = openpyxl.Workbook()
    wb.remove(wb["Sheet"])
    w_sheet = wb.create_sheet(title="相似度")
    base_cluster = [[5, 6],
                    [1, 14, 15],
                    [0, 2, 3, 4, 11, 12, 13],
                    [7, 8, 9, 10, 16, 17]]
    for k in range(1000):
        cluster_centers = get_centroids(data_set, 4)
        final_cluster_center, final_cluster = generate_clusters(4, data_set, cluster_centers, len(data_set[0]))
        w_sheet.cell(row=k+1, column=1).value = get_similarity(base_cluster, final_cluster)
        base_cluster = copy.deepcopy(final_cluster)
    wb.save('PbBa-1000次实验-与相邻基准的相似度.xlsx')
